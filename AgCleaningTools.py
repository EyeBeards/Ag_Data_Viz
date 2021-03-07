from openpyxl import load_workbook
import pandas as pd
import datetime
import os

DATASETS = []

DATAFRAMES = {}


class Dataset():
    def __init__(self, _data):
        self.wb = load_workbook(_data, read_only=True)
        _month = int(_data[-11:-9])
        _day = int(_data[-9:-7])
        _year = int(_data[-7:-5])
        _date = datetime.date(2000+_year, _month, _day)
        print("\n\n{}:\t{}".format(_data, _date))

        self.data = {_data: {
                        "date": _date,
                        "sheets": []
                        }
                    }

        for _sheet in self.wb.sheetnames:
            self.data[_data]["sheets"].append(_sheet)
            df = pd.DataFrame(self.wb[_sheet].values)         

            if _sheet == "crop_summary" or _sheet == "county_summary" or _sheet == "prevent" or _sheet == "plant_and_fail":
                name_row = 3
            
            else:
                name_row = 0

            df.rename(columns=df.iloc[name_row], inplace = True)
            df = df.drop(name_row, axis=0)
            df.insert(0, "date", _date, True)
            print("SHEET: {}".format(_sheet))
            # print("DATAFRAME: \n{}".format(df))

            if _sheet in ["plant_and_fail", "prevent"]:
                end_row = 55
            elif _sheet == "county_summary":
                end_row = 3120
            elif _sheet == "crop_summary":
                end_row = 270
            else:
                end_row = None

            if _sheet not in list(DATAFRAMES.keys()):
                DATAFRAMES[_sheet] = pd.DataFrame(df)

            else:
                # end_row = None
                if end_row:
                    df = df.loc[name_row:end_row,:]

                else:
                    _df = df.loc[name_row:,:]

            print("DATAFRAME: \n{}".format(DATAFRAMES[_sheet]))
            print("df: \n{}".format(df))
                # DATAFRAMES[_sheet] = pd.concat([DATAFRAMES[_sheet], _df])
            

def main():
    for _file in os.popen("ls").read().split("\n"):
        
        if ".xlsx" in _file:
           DATASETS.append(Dataset(_file))

if __name__=="__main__":
    main()

